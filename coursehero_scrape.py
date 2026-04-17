#!/usr/bin/env python3
"""
CourseHero sitemap scraper — one script, many URLs.

First run (log in once):
    python3 coursehero_scrape.py --login

Scrape:
    python3 coursehero_scrape.py "URL1,URL2,URL3"
    python3 coursehero_scrape.py URL1 URL2 URL3 --out ./my_output_dir

Profile is persisted at ~/.coursehero-profile so login is reused.
Output: <out_dir>/coursehero_<date>.json and coursehero_<date>.xlsx.
"""

import argparse
import asyncio
import json
import math
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from playwright.async_api import async_playwright

PROFILE_DIR      = Path.home() / ".coursehero-profile"
DATE_STAMP       = datetime.now().strftime("%Y-%m-%d")
WAIT_AFTER_NAV   = 9
WAIT_AFTER_CLICK = 6
BETWEEN_PAGES    = 2


# ---------------------------------------------------------------------------
# Extraction helpers
# ---------------------------------------------------------------------------
def parse_pagination(html):
    m = re.search(r'Showing\s+(\d+)\s+to\s+(\d+)\s+of\s+(\d[\d,]*)', html)
    if m:
        start = int(m.group(1))
        end   = int(m.group(2))
        total = int(m.group(3).replace(',', ''))
        return total, max(end - start + 1, 1)
    soup = BeautifulSoup(html, 'lxml')
    p = soup.find('p', class_='tl_resourceContent_title')
    if p:
        m2 = re.search(r'of\s+(\d[\d,]*)', p.get_text())
        if m2:
            return int(m2.group(1).replace(',', '')), 30
    return 0, 30


def _parse_semester(text):
    if not text:
        return ''
    t = str(text)
    m = re.search(r'\b(Fall|Spring|Winter|Summer|Sp|Fl|Wi|Su)\s*(20\d{2}|1[89]\d{2})\b', t, re.I)
    if m:
        sem_map = {'fall':'Fall','fl':'Fall','spring':'Spring','sp':'Spring',
                   'winter':'Winter','wi':'Winter','summer':'Summer','su':'Summer'}
        return f"{sem_map.get(m.group(1).lower(), m.group(1).title())} {m.group(2)}"
    m = re.search(r'\b([FfWwSs])(\d{2})\b', t)
    if m:
        pm = {'f':'Fall','w':'Winter','s':'Spring'}
        return f"{pm.get(m.group(1).lower(),'?')} 20{m.group(2)}"
    m = re.search(r'\b(20\d{2}|19\d{2})\b', t)
    return m.group(1) if m else ''


def _get_ext(text):
    if not text:
        return ''
    m = re.search(r'\.(pdf|docx?|pptx?|xlsx?|txt|png|jpg|jpeg|gif|zip)(?:$|[^a-z])', text, re.I)
    return m.group(1).lower() if m else ''


def parse_school_and_course(source_url):
    m = re.search(r'/schools/\d+-([^/]+)/courses/\d+-([^/]+)', source_url)
    if not m:
        return 'Unknown School', 'UNKNOWN'
    school = m.group(1).replace('-', ' ')
    course = m.group(2)
    return school, course


_STOP_WORDS = {'of', 'the', 'and', 'at', 'for', 'a', 'an', 'in', 'on'}

def _school_initials(source_url):
    """Derive initials from a school slug in a sitemap URL. Falls back to 'SCHL'."""
    m = re.search(r'/schools/\d+-([^/]+)/', source_url)
    if not m:
        return 'SCHL'
    parts = [p for p in m.group(1).split('-') if p]
    letters = [p[0] for p in parts if p.lower() not in _STOP_WORDS]
    return ''.join(letters).upper() or 'SCHL'


def school_tag(urls):
    """Combine unique school initials from a list of URLs into a single tag."""
    seen = []
    for u in urls:
        ini = _school_initials(u)
        if ini not in seen:
            seen.append(ini)
    return '_'.join(seen) if seen else 'SCHL'


def extract_documents(html, source_url, page_num):
    """
    Pull every document-like card off a sitemap page. No subject-specific rules.
    Handles both the new Tailwind-based layout and the old legacy layout.
    """
    soup = BeautifulSoup(html, 'lxml')
    inferred_school, inferred_course = parse_school_and_course(source_url)

    new_items = soup.find_all('li', attrs={'aria-label': re.compile(r'^(documents|trending)-\d+')})
    old_items = soup.find_all('li', class_='tl_documents_list-item')

    results = []
    seen    = set()

    # ---- NEW Tailwind layout ----
    for item in new_items:
        a_tag = item.find('a', href=lambda h: h and '/file/' in h)
        if a_tag:
            href = a_tag['href']
            if href.startswith('/'):
                href = 'https://www.coursehero.com' + href
            file_url = href.rstrip('/')
        else:
            if not item.find('a', href=True):
                continue
            file_url = '(not available)'

        if file_url != '(not available)':
            if file_url in seen:
                continue
            seen.add(file_url)

        h3    = item.find('h3')
        title = h3.get_text(strip=True) if h3 else ''
        if not title and a_tag and a_tag.get('title'):
            title = a_tag['title']

        # Skip items with neither a title nor a real file URL — extraction noise.
        if not title and file_url == '(not available)':
            continue

        footer      = item.find('footer')
        course_code = ''
        school      = inferred_school
        if footer:
            h4 = footer.find('h4')
            if h4:
                course_code = h4.get_text(strip=True)
            school_div = footer.find('div', class_=re.compile(r'tw-truncate'))
            if school_div:
                school = school_div.get_text(strip=True) or school

        semester_year = _parse_semester(title) or _parse_semester(file_url)
        file_ext      = _get_ext(title) or _get_ext(file_url) or 'pdf'

        pages     = ''
        page_span = item.find('span', string=re.compile(r'\d+\s+page', re.I))
        if page_span:
            pm = re.search(r'(\d+)', page_span.get_text())
            if pm:
                pages = pm.group(1)

        desc_p      = item.select_one('p[class*="tw-line-clamp"]')
        description = desc_p.get_text('\n', strip=True) if desc_p else ''

        results.append({
            'title':         title,
            'url':           file_url,
            'course_code':   course_code or inferred_course,
            'school':        school,
            'semester_year': semester_year,
            'pages':         pages,
            'file_ext':      file_ext,
            'description':   description,
            'source_url':    source_url,
            'page_scraped':  page_num,
        })

    # ---- OLD legacy layout (fallback) ----
    for item in old_items:
        file_url = ''
        for a in item.find_all('a', href=True):
            href = a['href']
            if href.startswith('/file/'):
                file_url = 'https://www.coursehero.com' + href.rstrip('/')
                break
        if not file_url:
            reg_a = item.find('a', href=re.compile(r'/register/\?get_doc='))
            if reg_a:
                m = re.search(r'get_doc=(\d+)', reg_a['href'])
                if m:
                    file_url = f'https://www.coursehero.com/file/{m.group(1)}/'
        if not file_url:
            file_url = '(not available)'

        if file_url != '(not available)':
            if file_url in seen:
                continue
            seen.add(file_url)

        title_li = item.find('li', class_=re.compile(r'ch_product_document_title'))
        title    = title_li.get_text(strip=True) if title_li else ''
        if not title:
            footer_div = item.find('div', class_='ch_product_document_footer')
            title      = footer_div.get_text(strip=True) if footer_div else ''

        pages_span = item.find('span', class_='ch_product_document_count')
        pages      = pages_span.get_text(strip=True) if pages_span else ''

        school_li = item.find('li', class_='ch_product_document_meta-school')
        school    = school_li.get_text(strip=True) if school_li else inferred_school

        meta_li       = item.find('li', class_='meta-course_nosnippet')
        course_code   = ''
        semester_year = ''
        if meta_li:
            raw = re.sub(r'\s+', ' ', meta_li.get_text()).strip()
            mm  = re.match(r'([A-Z]+\s*\d+[A-Z]*)\s*[-–]\s*(.*)', raw)
            if mm:
                course_code   = mm.group(1).strip()
                semester_year = mm.group(2).strip()
            else:
                course_code = raw

        file_ext = _get_ext(title) or _get_ext(file_url) or 'pdf'

        results.append({
            'title':         title,
            'url':           file_url,
            'course_code':   course_code or inferred_course,
            'school':        school,
            'semester_year': semester_year,
            'pages':         pages,
            'file_ext':      file_ext,
            'description':   '',
            'source_url':    source_url,
            'page_scraped':  page_num,
        })

    return results


# ---------------------------------------------------------------------------
# Generic document-type classifier — no subject-specific keywords
# ---------------------------------------------------------------------------
TYPE_ORDER = [
    'Final Exam', 'Midterm', 'Exam', 'Test',
    'Practice Exam / Sample Exam', 'Quiz',
    'Lab Report / Lab Manual', 'Study Guide / Review Sheet',
    'Syllabus', 'Homework / Problem Set', 'Solutions / Answer Key',
    'Notes / Lecture Slides', 'Handout / Worksheet',
    'Formula Sheet / Cheat Sheet', 'Summary / Outline',
    'Textbook Chapter / Reading', 'Screenshot / Image',
    'Course Schedule', 'Not Available', 'Other / Miscellaneous',
]


def classify_doc_type(name, url=''):
    if url == '(not available)' or (name and name.startswith('[Page ')):
        return 'Not Available'
    t = (name or '').lower()
    if re.search(r'practice\s*(exam|final|midterm|test|quiz)|sample\s*(exam|test)|mock\s*exam|practice\s*problems?', t):
        return 'Practice Exam / Sample Exam'
    if re.search(r'final\s*exam|exam\s*final', t):
        return 'Final Exam'
    if re.search(r'\bfinal\b', t) and re.search(r'\bkey\b|\banswers?\b|\bsolution', t):
        return 'Final Exam'
    if re.search(r'\bmidterm\b|\bmid[\s\-]?term\b', t):
        return 'Midterm'
    if re.search(r'\bexam\s*\d+|\bexam\b', t):
        return 'Exam'
    if re.search(r'\btest\s*\d+|\btest\b', t):
        return 'Test'
    if re.search(r'\bquiz\b', t):
        return 'Quiz'
    if re.search(r'\blab\s*(report|manual|notebook|experiment|data|discussion|activity|quiz)|\bpre[\s\-]?lab|\bpost[\s\-]?lab|\bexperiment\s*(report|write[\s\-]?up|procedure)', t):
        return 'Lab Report / Lab Manual'
    if re.search(r'\blab\s*\d+|\bexperiment\s*\d+', t):
        return 'Lab Report / Lab Manual'
    if re.search(r'study\s*guide|review\s*sheet|exam\s*review|test\s*review|midterm\s*review|final\s*review|study\s*outline|review\s*guide', t):
        return 'Study Guide / Review Sheet'
    if re.search(r'\bsyllabus\b', t):
        return 'Syllabus'
    if re.search(r'\bhomework\b|\bhw\s*\d|\bproblem\s*set[s]?\b|\bpset\s*\d|\bps\s*\d\b', t):
        return 'Homework / Problem Set'
    if re.search(r'\bsolution[s]?\b|\banswer\s*key\b', t):
        return 'Solutions / Answer Key'
    if re.search(r'\bnote[s]?\b|\blecture\b|\bslide[s]?\b|\bppt\b|\bpowerpoint\b', t):
        return 'Notes / Lecture Slides'
    if re.search(r'\bhandout\b|\bworksheet\b', t):
        return 'Handout / Worksheet'
    if re.search(r'\bformula\b|\bcheat\s*sheet\b|\bequation\s*sheet\b|\bcrib\b', t):
        return 'Formula Sheet / Cheat Sheet'
    if re.search(r'\bsummary\b|\boutline\b|\boverview\b', t):
        return 'Summary / Outline'
    if re.search(r'\bchapter\s*\d+|\btextbook\b|\breading\b', t):
        return 'Textbook Chapter / Reading'
    if re.search(r'\bscreenshot\b|\bimage\b|\bimg\b', t):
        return 'Screenshot / Image'
    if re.search(r'\bschedul|\bcalendar\b', t):
        return 'Course Schedule'
    return 'Other / Miscellaneous'


def has_key(name):
    return bool(re.search(r'\bkey\b|\banswers?\b|\bsolution[s]?\b', (name or '').lower()))


def clean_title(title):
    if not title:
        return title
    return re.sub(r'^\d+\s+pages?\s*', '', title, flags=re.I).strip()


def slug_to_name(url):
    if not url or url == '(not available)':
        return url
    m = re.search(r'/file/\d+/(.+?)/?$', url)
    if not m:
        return url
    slug = m.group(1)
    slug = slug.replace('%20', ' ').replace('%2B', '+').replace('%28', '(').replace('%29', ')')
    slug = re.sub(r'(pdf|docx?|pptx?|xlsx?|txt|png|jpg|jpeg|gif|zip)$', r'.\1', slug, flags=re.I)
    slug = slug.replace('-', ' ').replace('_', ' ')
    return re.sub(r'\s+', ' ', slug).strip()


_ILLEGAL_XLSX = re.compile(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]')

def _sanitize(text):
    """Strip characters openpyxl rejects (C0 control chars except \\t \\n \\r)."""
    if not text:
        return ''
    return _ILLEGAL_XLSX.sub('', text)


def enrich(raw):
    url   = raw.get('url', '').strip()
    title = clean_title(raw.get('title', '')) or slug_to_name(url)
    sem   = raw.get('semester_year') or _parse_semester(title)
    return {
        'url':           url,
        'title':         _sanitize(title),
        'course_code':   raw.get('course_code', ''),
        'school':        raw.get('school', ''),
        'semester_year': sem,
        'pages':         raw.get('pages', ''),
        'file_ext':      raw.get('file_ext', 'pdf'),
        'description':   _sanitize(raw.get('description', '')),
        'source_url':    raw.get('source_url', ''),
        'doc_type':      classify_doc_type(title, url),
        'has_key':       has_key(title) if url != '(not available)' else False,
    }


# ---------------------------------------------------------------------------
# Playwright scraping
# ---------------------------------------------------------------------------
async def launch_context(pw, headless=False):
    PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    return await pw.chromium.launch_persistent_context(
        str(PROFILE_DIR),
        headless=headless,
        viewport={'width': 1440, 'height': 900},
        args=['--disable-blink-features=AutomationControlled'],
    )


async def is_logged_in(ctx):
    cookies = await ctx.cookies('https://www.coursehero.com')
    wanted = {'cs_session_auth', 'CS_AUTHENTICATION_COOKIE', 'login_id',
              'cshero_session', 'ch_logged_in', 'userAuth', 'xsrf_token'}
    return any(c.get('name') in wanted for c in cookies)


async def interactive_login(ctx):
    page = await ctx.new_page()
    await page.goto('https://www.coursehero.com/login/', wait_until='domcontentloaded')
    print("\n" + "=" * 70)
    print("Log in to CourseHero in the browser window (Google SSO is fine).")
    print("Solve any captcha. Then return to this terminal.")
    print("=" * 70)
    input("Press ENTER after login is complete... ")
    await page.close()


async def scrape_one_url(ctx, source_url, out_dir, idx):
    print(f"\n{'='*70}")
    print(f"[{idx}] Scraping: {source_url}")
    page = await ctx.new_page()

    try:
        await page.goto(source_url, timeout=60000, wait_until='domcontentloaded')
    except Exception as e:
        print(f"  ERROR navigating: {e}")
        await page.close()
        return [], {'status': 'error', 'error': str(e)[:120]}

    print(f"  Waiting {WAIT_AFTER_NAV}s for JS render...")
    await asyncio.sleep(WAIT_AFTER_NAV)

    html = await page.content()
    safe = re.sub(r'[^\w]+', '_', source_url).strip('_')[-80:]
    (out_dir / f'raw_{safe}_p1.html').write_text(html, encoding='utf-8')

    status = 'success'
    err    = ''
    if 'cloudflare' in html.lower() and 'checking' in html.lower():
        status = 'cloudflare'
        err    = 'Cloudflare challenge page detected'
    elif 'log in' in html.lower() and 'sitemap' not in html.lower():
        status = 'login_required'
        err    = 'Login required — profile may have expired'

    total, per_page = parse_pagination(html)
    print(f"  Total documents: {total} ({per_page} per page)")

    all_docs  = []
    seen_urls = set()

    docs = extract_documents(html, source_url, 1)
    for d in docs:
        key = d['url']
        if key != '(not available)' and key in seen_urls:
            continue
        if key != '(not available)':
            seen_urls.add(key)
        all_docs.append(d)
    print(f"  Page 1: {len(docs)} docs (running {len(all_docs)})")

    if total == 0 and not docs:
        await page.close()
        return all_docs, {'status': status or 'success', 'error': err}

    total_pages = math.ceil(total / per_page) if per_page > 0 else 1
    page_num    = 1

    while page_num < total_pages:
        page_num += 1
        print(f"  Page {page_num}/{total_pages} — clicking Next...")
        await asyncio.sleep(BETWEEN_PAGES)

        clicked = await page.evaluate("""
            () => {
                const items = Array.from(document.querySelectorAll('li'));
                const btn = items.find(el => {
                    const span = el.querySelector('span');
                    const txt = span ? span.textContent.trim() : '';
                    return txt === 'Next' && (el.className || '').includes('cursor-pointer');
                });
                if (btn) { btn.click(); return true; }
                return false;
            }
        """)

        if not clicked:
            print(f"  Next button not found/disabled — stopping at page {page_num - 1}.")
            break

        await asyncio.sleep(WAIT_AFTER_CLICK)
        html_n = await page.content()
        (out_dir / f'raw_{safe}_p{page_num}.html').write_text(html_n, encoding='utf-8')

        docs_n = extract_documents(html_n, source_url, page_num)
        added  = 0
        for d in docs_n:
            key = d['url']
            if key != '(not available)' and key in seen_urls:
                continue
            if key != '(not available)':
                seen_urls.add(key)
            all_docs.append(d)
            added += 1
        print(f"  Page {page_num}: {len(docs_n)} extracted, {added} new (running {len(all_docs)})")

        if not docs_n:
            break

    await page.close()
    return all_docs, {'status': status or 'success', 'error': err}


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------
def write_xlsx(docs, out_path, url_status, school_label):
    wb = Workbook()

    HDR_BLUE   = "1F4E79"
    HDR_GRAY   = "666666"
    ALT_ROW    = "E8EEF5"
    LIGHT_BG   = "F0F4F9"
    NA_BG      = "F0F0F0"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=HDR_BLUE, end_color=HDR_BLUE, fill_type="solid")
    header_gray = PatternFill(start_color=HDR_GRAY, end_color=HDR_GRAY, fill_type="solid")
    wrap_align  = Alignment(vertical="top", wrap_text=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),  right=Side(style='thin'),
        top=Side(style='thin'),   bottom=Side(style='thin'),
    )
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_alt   = PatternFill(start_color=ALT_ROW,  end_color=ALT_ROW,  fill_type="solid")
    fill_na    = PatternFill(start_color=NA_BG,    end_color=NA_BG,    fill_type="solid")

    def style_header(ws, headers, gray=False):
        ws.append(headers)
        fill = header_gray if gray else header_fill
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = fill
            cell.alignment = header_align
            cell.border    = thin_border
        ws.row_dimensions[1].height = 32

    def style_row(ws, row_num, is_na=False):
        if is_na:
            row_fill   = fill_na
            fnt_color  = "999999"
            fnt_italic = True
        else:
            row_fill   = fill_alt if row_num % 2 == 0 else fill_white
            fnt_color  = "000000"
            fnt_italic = False
        for cell in ws[ws.max_row]:
            cell.font      = Font(color=fnt_color, size=10, italic=fnt_italic)
            cell.fill      = row_fill
            cell.alignment = wrap_align
            cell.border    = thin_border

    def type_sort_key(d):
        try:
            ti = TYPE_ORDER.index(d['doc_type'])
        except ValueError:
            ti = len(TYPE_ORDER)
        return (ti, d['semester_year'] or 'zzzz', d['title'].lower())

    docs_sorted = sorted(docs, key=type_sort_key)

    # ── Sheet: All Documents ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "All Documents"
    headers = ["#", "Course Code", "School", "Document Name", "Document Type",
               "Semester / Year", "Has Key?", "File Type", "Pages", "Description", "URL"]
    style_header(ws, headers)
    for i, d in enumerate(docs_sorted, 1):
        is_na = d['url'] == '(not available)'
        ws.append([
            i, d['course_code'], d['school'], d['title'], d['doc_type'],
            d['semester_year'], 'Yes' if d['has_key'] else '',
            d['file_ext'].upper() if d['file_ext'] and not is_na else '',
            d['pages'], d.get('description', ''), d['url'],
        ])
        style_row(ws, i, is_na=is_na)
    ws.freeze_panes       = "A2"
    ws.auto_filter.ref    = ws.dimensions
    for col, w in zip('ABCDEFGHIJK', [5, 16, 26, 60, 28, 16, 10, 10, 8, 60, 72]):
        ws.column_dimensions[col].width = w

    # ── Sheet: By Document Type ─────────────────────────────────────────────
    ws_t = wb.create_sheet("By Document Type")
    style_header(ws_t, ["#", "Document Type", "Course Code", "Document Name",
                        "Semester / Year", "Has Key?", "File Type", "Description", "URL"], gray=True)
    for i, d in enumerate(docs_sorted, 1):
        is_na = d['url'] == '(not available)'
        ws_t.append([
            i, d['doc_type'], d['course_code'], d['title'],
            d['semester_year'], 'Yes' if d['has_key'] else '',
            d['file_ext'].upper() if d['file_ext'] and not is_na else '',
            d.get('description', ''), d['url'],
        ])
        style_row(ws_t, i, is_na=is_na)
    ws_t.freeze_panes    = "A2"
    ws_t.auto_filter.ref = ws_t.dimensions
    for col, w in zip('ABCDEFGHI', [5, 30, 16, 60, 16, 10, 10, 60, 72]):
        ws_t.column_dimensions[col].width = w

    # ── Sheet: By Course ────────────────────────────────────────────────────
    def course_sort_key(d):
        return (d['course_code'] or 'zzz',) + type_sort_key(d)

    ws_c = wb.create_sheet("By Course")
    style_header(ws_c, ["#", "Course Code", "Document Name", "Document Type",
                        "Semester / Year", "Has Key?", "File Type", "Pages", "Description", "URL"])
    for i, d in enumerate(sorted(docs, key=course_sort_key), 1):
        is_na = d['url'] == '(not available)'
        ws_c.append([
            i, d['course_code'], d['title'], d['doc_type'],
            d['semester_year'], 'Yes' if d['has_key'] else '',
            d['file_ext'].upper() if d['file_ext'] and not is_na else '',
            d['pages'], d.get('description', ''), d['url'],
        ])
        style_row(ws_c, i, is_na=is_na)
    ws_c.freeze_panes    = "A2"
    ws_c.auto_filter.ref = ws_c.dimensions
    for col, w in zip('ABCDEFGHIJ', [5, 16, 60, 28, 16, 10, 10, 8, 60, 72]):
        ws_c.column_dimensions[col].width = w

    # ── Sheet: Summary ──────────────────────────────────────────────────────
    ws_s = wb.create_sheet("Summary")
    ws_s.column_dimensions['A'].width = 52
    ws_s.column_dimensions['B'].width = 80

    def srow(label, value='', bold=False, fill=None, font_color="000000"):
        ws_s.append([str(label), str(value) if value != '' else ''])
        r = ws_s.max_row
        ws_s[f'A{r}'].font      = Font(bold=bold, size=10, color=font_color)
        ws_s[f'B{r}'].font      = Font(size=10, color=font_color)
        ws_s[f'A{r}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws_s[f'B{r}'].alignment = Alignment(wrap_text=True, vertical='top')
        if fill:
            f = PatternFill(start_color=fill, end_color=fill, fill_type='solid')
            ws_s[f'A{r}'].fill = f
            ws_s[f'B{r}'].fill = f

    srow(f"{school_label} — CourseHero Document Inventory", "",
         bold=True, fill=HDR_BLUE, font_color="FFFFFF")
    ws_s.row_dimensions[1].height = 28
    ws_s.append([])

    total_ok = sum(1 for d in docs if d['url'] != '(not available)')
    total_na = sum(1 for d in docs if d['url'] == '(not available)')
    srow("Search Date",           DATE_STAMP, bold=True)
    srow("Total Documents",       len(docs),  bold=True)
    srow("  — Available URLs",    total_ok)
    srow("  — Not Available",     total_na)
    ws_s.append([])

    course_counts = Counter(d['course_code'] for d in docs if d['course_code'])
    if course_counts:
        srow("Documents per Course Code", "", bold=True, fill=LIGHT_BG)
        ws_s.append(["Course Code", "Count"])
        r = ws_s.max_row
        ws_s[f'A{r}'].font = Font(bold=True, size=10)
        ws_s[f'B{r}'].font = Font(bold=True, size=10)
        for code, ct in course_counts.most_common():
            srow(code, ct)
        ws_s.append([])

    srow("Source URLs + Scrape Status", "", bold=True, fill=LIGHT_BG)
    ws_s.append(["Status / URL", "URL"])
    r = ws_s.max_row
    for col in ('A', 'B'):
        ws_s[f'{col}{r}'].font = Font(bold=True, size=10, color="FFFFFF")
        ws_s[f'{col}{r}'].fill = PatternFill(start_color=HDR_BLUE, end_color=HDR_BLUE, fill_type="solid")
    for url, info in url_status.items():
        st  = info.get('status', 'unknown')
        err = info.get('error', '')
        if st == 'success':
            label = "SUCCESS — fully scraped"; fc = "006400"
        elif st == 'cloudflare':
            label = "NOT SCRAPED — Cloudflare block"; fc = "CC0000"
        elif st == 'timeout':
            label = "NOT SCRAPED — timeout"; fc = "CC6600"
        elif st == 'login_required':
            label = "NOT SCRAPED — login required"; fc = "CC0000"
        elif st == 'error':
            label = f"ERROR — {err[:50]}"; fc = "CC0000"
        else:
            label = f"UNKNOWN ({st})"; fc = "666666"
        ws_s.append([label, url])
        rr = ws_s.max_row
        ws_s[f'A{rr}'].font = Font(bold=True, size=10, color=fc)
        ws_s[f'B{rr}'].font = Font(size=10, color="0563C1", underline='single')
        ws_s[f'A{rr}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws_s[f'B{rr}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws_s.append([])

    type_counts = Counter(d['doc_type'] for d in docs)
    srow("Documents by Type", "", bold=True, fill=LIGHT_BG)
    ws_s.append(["Document Type", "Count"])
    r = ws_s.max_row
    ws_s[f'A{r}'].font = Font(bold=True, size=10)
    ws_s[f'B{r}'].font = Font(bold=True, size=10)
    for t in TYPE_ORDER:
        if t in type_counts:
            srow(t, type_counts[t])
    for t, ct in type_counts.items():
        if t not in TYPE_ORDER:
            srow(t, ct)

    wb.save(out_path)
    return type_counts


# ---------------------------------------------------------------------------
# Main orchestration
# ---------------------------------------------------------------------------
async def run(urls, out_dir, headless, tag):
    out_dir.mkdir(parents=True, exist_ok=True)
    raw_dir = out_dir / 'raw_html'
    raw_dir.mkdir(exist_ok=True)

    async with async_playwright() as pw:
        ctx = await launch_context(pw, headless=headless)

        if not await is_logged_in(ctx):
            print("Not logged in — opening login page.")
            await interactive_login(ctx)

        all_raw    = []
        url_status = {}
        seen_global = set()

        for idx, url in enumerate(urls, 1):
            try:
                docs, status = await scrape_one_url(ctx, url, raw_dir, idx)
            except Exception as e:
                print(f"  FATAL: {e}")
                docs   = []
                status = {'status': 'error', 'error': str(e)[:120]}

            url_status[url] = status
            for d in docs:
                key = d['url']
                if key != '(not available)' and key in seen_global:
                    continue
                if key != '(not available)':
                    seen_global.add(key)
                all_raw.append(d)

            # checkpoint after each URL
            (out_dir / f'checkpoint_{tag}_{DATE_STAMP}.json').write_text(
                json.dumps({'documents': all_raw, 'url_status': url_status},
                           indent=2, ensure_ascii=False), encoding='utf-8'
            )

        await ctx.close()

    # Enrich + write outputs
    docs_enriched = [enrich(d) for d in all_raw]

    json_path = out_dir / f'coursehero_{tag}_{DATE_STAMP}.json'
    json_path.write_text(
        json.dumps({'documents': docs_enriched, 'url_status': url_status},
                   indent=2, ensure_ascii=False), encoding='utf-8'
    )

    # school label: most common school across docs
    schools = Counter(d['school'] for d in docs_enriched if d['school'])
    school_label = schools.most_common(1)[0][0] if schools else 'CourseHero'

    xlsx_path = out_dir / f'coursehero_{tag}_{DATE_STAMP}.xlsx'
    type_counts = write_xlsx(docs_enriched, xlsx_path, url_status, school_label)

    print(f"\n{'='*70}")
    print(f"DONE — {len(docs_enriched)} documents")
    print(f"  JSON : {json_path}")
    print(f"  XLSX : {xlsx_path}")
    print("\nTypes:")
    for t in TYPE_ORDER:
        if t in type_counts:
            print(f"  {t}: {type_counts[t]}")

    print("\nURL status:")
    for url, info in url_status.items():
        mark = "✓" if info['status'] == 'success' else "✗"
        print(f"  {mark} {info['status']}: {url}")


async def login_only():
    async with async_playwright() as pw:
        ctx = await launch_context(pw, headless=False)
        if await is_logged_in(ctx):
            print("Already logged in — profile is ready.")
        else:
            await interactive_login(ctx)
            print("Login saved to profile at:", PROFILE_DIR)
        await ctx.close()


def parse_args():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('urls', nargs='*',
                    help='CourseHero sitemap URLs (space- or comma-separated)')
    ap.add_argument('--out', type=Path, default=None,
                    help='Output directory (default: ./coursehero_<SCHOOL>_<date>/)')
    ap.add_argument('--headless', action='store_true',
                    help='Run browser headless (not recommended for login)')
    ap.add_argument('--login', action='store_true',
                    help='Log in interactively and exit (first-time setup)')
    return ap.parse_args()


def main():
    args = parse_args()

    if args.login:
        asyncio.run(login_only())
        return

    raw = ' '.join(args.urls) if args.urls else ''
    if not raw.strip():
        print("ERROR: No URLs provided.")
        print("  Run with --login once, then pass URLs:")
        print('  python3 coursehero_scrape.py "URL1,URL2,URL3"')
        sys.exit(1)

    urls = [u.strip() for u in re.split(r'[,\s]+', raw) if u.strip()]
    tag  = school_tag(urls)
    out  = args.out or Path(f'./coursehero_{tag}_{DATE_STAMP}')

    if out.exists() and any(out.iterdir()):
        print(f"WARNING: output directory {out} exists and is not empty — files may be overwritten.")

    print(f"Will scrape {len(urls)} URL(s) → {out}")
    asyncio.run(run(urls, out, args.headless, tag))


if __name__ == '__main__':
    main()
