#!/usr/bin/env python3
"""
Generate comprehensive XLSX inventory of all CourseHero documents for
University of California Santa Barbara — Chemistry 1C
across CHEM1C, CHEM1CL, CHEM-1C1C, and CHEMISTRY1C course codes.

Data sourced from scraping the following sitemap URLs (paginated):
  CHEM 1C:       https://www.coursehero.com/sitemap/schools/27-University-of-California-Santa-Barbara/courses/242694-CHEM1C/
  CHEM 1CL:      https://www.coursehero.com/sitemap/schools/27-University-of-California-Santa-Barbara/courses/242693-CHEM1CL/
  CHEM-1C1C:     https://www.coursehero.com/sitemap/schools/27-University-of-California-Santa-Barbara/courses/906488-CHEM-1C1C/
  CHEMISTRY 1C:  https://www.coursehero.com/sitemap/schools/27-University-of-California-Santa-Barbara/courses/4576471-CHEMISTRY1C/

Usage:
    source .venv/bin/activate
    cd "COURSEHERO SEARCHES_CoPilot/Coursehero UCSB CHEM 1C"
    python3 generate_ucsb_chem1c_xlsx.py [--json ucsb_chem1c_docs_YYYY-MM-DD.json]
"""

import re
import json
import os
import sys
import glob
from datetime import datetime
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATE_STAMP = datetime.now().strftime('%Y-%m-%d')


# ---------------------------------------------------------------------------
# Locate the JSON data file
# ---------------------------------------------------------------------------
def find_json_file():
    for i, arg in enumerate(sys.argv):
        if arg == '--json' and i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    patterns = [
        os.path.join(SCRIPT_DIR, 'ucsb_chem1c_docs_*.json'),
    ]
    candidates = []
    for p in patterns:
        candidates.extend(glob.glob(p))
    non_cp = [f for f in candidates if 'checkpoint' not in f]
    if non_cp:
        return max(non_cp, key=os.path.getmtime)
    if candidates:
        return max(candidates, key=os.path.getmtime)
    return None


# ---------------------------------------------------------------------------
# Title helpers
# ---------------------------------------------------------------------------
def clean_title(title):
    if not title:
        return title
    title = re.sub(r'^\d+\s+pages?\s*', '', title, flags=re.I).strip()
    return title


def slug_to_name(url):
    m = re.search(r'/file/\d+/(.+?)/?$', url)
    if not m:
        return url
    slug = m.group(1)
    slug = slug.replace('%20', ' ').replace('%2B', '+').replace('%28', '(').replace('%29', ')')
    slug = re.sub(r'(pdf|docx?|pptx?|xlsx?|txt|png|jpg|jpeg|gif|zip)$', r'.\1', slug, flags=re.I)
    slug = slug.replace('-', ' ').replace('_', ' ')
    slug = re.sub(r'\s+', ' ', slug).strip()
    return slug


# ---------------------------------------------------------------------------
# Document type classifier (Chemistry-focused)
# ---------------------------------------------------------------------------
TYPE_ORDER = [
    'Final Exam',
    'Midterm',
    'Exam',
    'Test',
    'Practice Exam / Sample Exam',
    'Quiz',
    'Lab Report / Lab Manual',
    'Study Guide / Review Sheet',
    'Syllabus',
    'Homework / Problem Set',
    'Solutions / Answer Key',
    'Notes / Lecture Slides',
    'Handout / Worksheet',
    'Formula Sheet / Cheat Sheet',
    'Summary / Outline',
    'Textbook Chapter / Reading',
    'Screenshot / Image',
    'Course Schedule',
    'Instructional Material',
    'Other / Miscellaneous',
]


def classify_doc_type(name):
    t = name.lower()

    # Practice / sample / mock exams
    if re.search(r'practice\s*(exam|final|midterm|test|quiz)|sample\s*(exam|test)|mock\s*exam|practice\s*problems?', t):
        return 'Practice Exam / Sample Exam'

    # Final exam
    if re.search(r'final\s*exam|exam\s*final', t):
        return 'Final Exam'
    if re.search(r'\bfinal\b', t) and re.search(r'\bkey\b|\banswers?\b|\bsolution', t):
        return 'Final Exam'
    if re.search(r'\bfinal\b', t) and not re.search(r'review|study|summary|note|prep|guide|outline|project|report|lab', t):
        return 'Final Exam'

    # Midterms
    if re.search(r'\bmidterm\b|\bmid[\s\-]?term\b|\bmid[\s\-]?exam\b', t):
        return 'Midterm'

    # Exams (numbered or general)
    if re.search(r'\bexam\s*[1-9ivxi]+\b|\b[1-9]\s*exam\b', t):
        return 'Exam'
    if re.search(r'\bexam\b', t) and not re.search(r'practice|sample|mock', t):
        return 'Exam'

    # Tests
    if re.search(r'\btest\s*[1-9ivxi]+\b|\btest\b', t) and re.search(r'key|answer|solution|blank', t):
        return 'Test'
    if re.search(r'\btest\s*[1-9]\b', t):
        return 'Test'

    # Quizzes
    if re.search(r'\bquiz\s*\d*\b|\bquiz\b', t):
        return 'Quiz'

    # Lab reports / manuals
    if re.search(r'\blab\s*(report|manual|notebook|experiment|data|discussion|activity|quiz|pre[\s\-]?lab|post[\s\-]?lab)\b|\bpre[\s\-]?lab\b|\bpost[\s\-]?lab\b', t):
        return 'Lab Report / Lab Manual'
    if re.search(r'\blab\s*\d+\b|\bexperiment\s*\d+\b', t):
        return 'Lab Report / Lab Manual'

    # Study guides / review
    if re.search(r'study\s*guide|review\s*sheet|exam\s*review|test\s*review|midterm\s*review|final\s*review|study\s*outline|review\s*guide|study\s*material|review\s*session|study\s*session', t):
        return 'Study Guide / Review Sheet'

    # Syllabus
    if re.search(r'\bsyllabus\b', t):
        return 'Syllabus'

    # Homework / problem sets
    if re.search(r'\bhomework\b|\bhw\s*\d|\bproblem\s*set[s]?\b|\bpset\s*\d|\bps\s*\d\b|\bproblem\s*set\b', t):
        return 'Homework / Problem Set'
    if re.search(r'\bhw\b', t) and re.search(r'\d|solution|key|answer', t):
        return 'Homework / Problem Set'

    # Solutions / answer keys
    if re.search(r'\bsolution[s]?\b|\banswer\s*key\b', t) and not re.search(r'exam|quiz|test|midterm|final', t):
        return 'Solutions / Answer Key'
    if re.search(r'\bkey\b', t) and re.search(r'hw|homework|problem\s*set|ps\d', t):
        return 'Solutions / Answer Key'

    # Textbook chapters
    if re.search(r'\bchapter\s*\d+\b|\bch\s*\d+\b|\btextbook\b|\bbook\b', t) and not re.search(r'problem|exercise|homework', t):
        return 'Textbook Chapter / Reading'
    if re.search(r'\bchapter\s*\d+\b', t) and re.search(r'problem|exercise|question|homework', t):
        return 'Homework / Problem Set'

    # Notes / lecture slides
    if re.search(r'\bnote[s]?\b|\blecture\b|\bslide[s]?\b|\bclass\s*note|\bpowerpoint\b|\bppt\b', t):
        return 'Notes / Lecture Slides'

    # Formula / cheat sheet
    if re.search(r'\bformula\b|\bcheat\s*sheet\b|\bcrib\b|\bequation\s*sheet\b|\bsummary\s*sheet\b|\bref(erence)?\s*sheet\b', t):
        return 'Formula Sheet / Cheat Sheet'

    # Summary / outline
    if re.search(r'\bsummary\b|\boutline\b|\boverview\b', t):
        return 'Summary / Outline'

    # Handouts / worksheets
    if re.search(r'\bhandout\b|\bworksheet\b', t):
        return 'Handout / Worksheet'

    # Screenshots / images
    if re.search(r'\bscreenshot\b|\bimg\b|\bimage\b', t, re.I):
        return 'Screenshot / Image'

    # Schedule
    if re.search(r'\bschedul\b|\bcalendar\b', t):
        return 'Course Schedule'

    # Chemistry topic keywords → instructional material
    if re.search(
        r'\batom\b|\batoms\b|\bmolecule\b|\bmolecular\b'
        r'|\bbond\b|\bbonding\b|\bcovalent\b|\bionic\b|\bhydrogen\b'
        r'|\borbital\b|\belectron\b|\bquantum\b|\bperiodic\b'
        r'|\bthermodynamics\b|\bentropy\b|\benthalpyb'
        r'|\bkinetics\b|\brate\b|\bequilibrium\b|\breaction\b'
        r'|\bacid\b|\bbase\b|\bph\b|\bredox\b|\boxidation\b|\breduction\b'
        r'|\bpolar\b|\bnonpolar\b|\bsolubility\b|\bsolution\b'
        r'|\bresonance\b|\bprobability\b|\bwavefunction\b|\borbital\b'
        r'|\bhydration\b|\bionization\b|\bvsepr\b|\bgeometry\b'
        r'|\bspecific\s*heat\b|\bheat\b|\btemperature\b|\bpressure\b'
        r'|\bvolume\b|\bgas\b|\bliquid\b|\bsolid\b|\bphase\b'
        r'|\bconcentration\b|\bmolarity\b|\bmolality\b|\bstoichiometry\b'
        r'|\bnuclear\b|\bradioacti\b|\bisotope\b|\bmass\s*spectrum\b',
        t
    ):
        return 'Instructional Material'

    return 'Other / Miscellaneous'


# ---------------------------------------------------------------------------
# Semester/year extraction
# ---------------------------------------------------------------------------
def extract_semester_year(text):
    if not text:
        return ''
    t = str(text)
    m = re.search(r'\b(Fall|Spring|Winter|Summer|Sp|Fl|Wi|Su)\s*(20\d{2}|1[89]\d{2})\b', t, re.I)
    if m:
        sem = m.group(1).lower()
        yr  = m.group(2)
        sem_map = {
            'fall': 'Fall', 'fl': 'Fall',
            'spring': 'Spring', 'sp': 'Spring',
            'winter': 'Winter', 'wi': 'Winter',
            'summer': 'Summer', 'su': 'Summer',
        }
        return f"{sem_map.get(sem, sem.title())} {yr}"
    m = re.search(r'\b([FfWwSs][Ii]?)(\d{2})\b', t)
    if m:
        p  = m.group(1)[0].lower()
        yr = '20' + m.group(2)
        pm = {'f': 'Fall', 'w': 'Winter', 's': 'Spring'}
        return f"{pm.get(p, '?')} {yr}"
    m = re.search(r'\b(20\d{2}|19\d{2})\b', t)
    if m:
        return m.group(1)
    return ''


# ---------------------------------------------------------------------------
# Answer key detection
# ---------------------------------------------------------------------------
def has_key(name):
    t = name.lower()
    return bool(re.search(r'\bkey\b|\banswers?\b|\bsolution[s]?\b', t))


# ---------------------------------------------------------------------------
# Professor inference (known UCSB Chemistry professors)
# ---------------------------------------------------------------------------
UCSB_CHEM_PROFS = {
    'leal':       'Prof. Leal',
    'guldi':      'Prof. Guldi',
    'gonen':      'Prof. Gonen',
    'hartshorn': 'Prof. Hartshorn',
    'petrich':    'Prof. Petrich',
    'meldrum':    'Prof. Meldrum',
    'turner':     'Prof. Turner',
    'brock':      'Prof. Brock',
    'garcia':     'Prof. Garcia',
    'martinez':   'Prof. Martinez',
    'schmidt':    'Prof. Schmidt',
    'keinan':     'Prof. Keinan',
    'schoenbaum': 'Prof. Schoenbaum',
    'maier':      'Prof. Maier',
    'blakemore':  'Prof. Blakemore',
    'chemla':     'Prof. Chemla',
    'harman':     'Prof. Harman',
    'jenkins':    'Prof. Jenkins',
}


def infer_professor(name):
    t = name.lower()
    found = []
    for key, val in UCSB_CHEM_PROFS.items():
        if key in t:
            found.append(val)
    return ', '.join(found) if found else ''


# ---------------------------------------------------------------------------
# Load JSON and build enriched document list
# ---------------------------------------------------------------------------
def build_docs(json_file):
    with open(json_file, encoding='utf-8') as f:
        raw = json.load(f)

    # Handle both old and new JSON formats
    if isinstance(raw, dict) and 'documents' in raw:
        documents = raw['documents']
    else:
        documents = raw

    docs      = []
    seen_urls = set()

    for item in documents:
        url = item.get('url', '').rstrip('/')
        if not url:
            continue
        if url in seen_urls:
            continue
        seen_urls.add(url)

        raw_title = item.get('title', '')
        title     = clean_title(raw_title) or slug_to_name(url)

        sem_yr = item.get('semester_year', '')
        if not sem_yr:
            sem_yr = extract_semester_year(item.get('course_code', ''))
        if not sem_yr:
            sem_yr = extract_semester_year(title)

        course_label = item.get('course_label', '')
        dept_label   = item.get('dept_label', '')

        doc = {
            'url':          url,
            'title':        title,
            'course_label': course_label,
            'dept_label':   dept_label,
            'course_name':  item.get('course_name', ''),
            'course_code':  item.get('course_code', ''),
            'semester_year': sem_yr,
            'pages':        item.get('pages', ''),
            'file_ext':     item.get('file_ext', 'pdf'),
            'school':       item.get('school', 'University of California Santa Barbara'),
            'doc_type':     classify_doc_type(title),
            'has_key':      has_key(title),
            'professor':    infer_professor(title),
        }
        docs.append(doc)

    return docs


def load_url_status(json_file):
    """Load URL status from JSON."""
    try:
        with open(json_file, encoding='utf-8') as f:
            raw = json.load(f)
        if isinstance(raw, dict) and 'url_status' in raw:
            return raw['url_status']
    except Exception:
        pass
    return {}


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------
def write_xlsx(docs, json_file, output_path, url_status):
    wb = Workbook()

    # UCSB color scheme: Gold #004B89 (Navy-ish), Gold #FFB81C
    UCSB_GOLD   = "FFB81C"
    UCSB_BLUE   = "004B89"
    UCSB_CREAM  = "F9F7F4"
    ALT_ROW     = "F5F3F0"

    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    header_font_blue  = Font(bold=True, color=UCSB_BLUE, size=11)
    header_fill_gold  = PatternFill(start_color=UCSB_GOLD,  end_color=UCSB_GOLD,  fill_type="solid")
    header_fill_blue  = PatternFill(start_color=UCSB_BLUE,  end_color=UCSB_BLUE,  fill_type="solid")
    header_fill_cream = PatternFill(start_color=UCSB_CREAM, end_color=UCSB_CREAM, fill_type="solid")
    text_font         = Font(color="000000", size=10)
    header_align      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align        = Alignment(vertical="top", wrap_text=True)
    thin_border       = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_alt   = PatternFill(start_color=ALT_ROW,  end_color=ALT_ROW,  fill_type="solid")

    def style_header(ws, headers, use_blue=False):
        ws.append(headers)
        fill = header_fill_blue if use_blue else header_fill_gold
        fnt  = header_font_white
        for cell in ws[1]:
            cell.font      = fnt
            cell.fill      = fill
            cell.alignment = header_align
            cell.border    = thin_border
        ws.row_dimensions[1].height = 32

    def style_data_row(ws, row_num):
        row_fill = fill_alt if row_num % 2 == 0 else fill_white
        for cell in ws[ws.max_row]:
            cell.font      = text_font
            cell.fill      = row_fill
            cell.alignment = wrap_align
            cell.border    = thin_border

    # ── Sort: type order → semester desc → title ──────────────────────────
    def sort_key(d):
        t = d['doc_type']
        try:
            ti = TYPE_ORDER.index(t)
        except ValueError:
            ti = len(TYPE_ORDER)
        yr = d['semester_year'] or 'zzzz'
        return (ti, yr, d['title'].lower())

    docs_sorted = sorted(docs, key=sort_key)

    # Department/course prefixes present
    DEPT_LABELS = ['CHEM', 'CHEMISTRY']

    # ── Sheet 1: All Documents ────────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "All Documents"
    headers_all = [
        "#", "Course Code", "Document Name", "Document Type",
        "Semester / Year", "Has Key?", "Professor",
        "File Type", "Pages", "URL", "Source Course Label",
    ]
    style_header(ws_all, headers_all, use_blue=False)

    for i, doc in enumerate(docs_sorted, 1):
        ws_all.append([
            i,
            doc['course_label'],
            doc['title'],
            doc['doc_type'],
            doc['semester_year'],
            'Yes' if doc['has_key'] else '',
            doc['professor'],
            doc['file_ext'].upper() if doc['file_ext'] else 'PDF',
            doc['pages'],
            doc['url'],
            doc['course_label'],
        ])
        style_data_row(ws_all, i)

    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = ws_all.dimensions
    for col, w in zip('ABCDEFGHIJK', [5, 14, 60, 28, 16, 10, 20, 10, 8, 72, 20]):
        ws_all.column_dimensions[col].width = w

    # ── Sheet: By Document Type ───────────────────────────────────────────
    ws_type = wb.create_sheet("By Document Type")
    headers_type = [
        "#", "Document Type", "Course Code", "Document Name",
        "Semester / Year", "Has Key?", "Professor", "File Type", "URL",
    ]
    style_header(ws_type, headers_type, use_blue=True)

    def type_sort_key(d):
        t = d['doc_type']
        try:
            ti = TYPE_ORDER.index(t)
        except ValueError:
            ti = len(TYPE_ORDER)
        yr = d['semester_year'] or 'zzzz'
        return (ti, yr)

    docs_by_type = sorted(docs, key=type_sort_key)
    for i, doc in enumerate(docs_by_type, 1):
        ws_type.append([
            i,
            doc['doc_type'],
            doc['course_label'],
            doc['title'],
            doc['semester_year'],
            'Yes' if doc['has_key'] else '',
            doc['professor'],
            doc['file_ext'].upper() if doc['file_ext'] else 'PDF',
            doc['url'],
        ])
        style_data_row(ws_type, i)

    ws_type.freeze_panes = "A2"
    ws_type.auto_filter.ref = ws_type.dimensions
    for col, w in zip('ABCDEFGHI', [5, 30, 14, 60, 16, 10, 20, 10, 72]):
        ws_type.column_dimensions[col].width = w

    # ── Sheet: Exams & Quizzes Only ───────────────────────────────────────
    EXAM_TYPES = {
        'Final Exam', 'Midterm', 'Exam', 'Test',
        'Practice Exam / Sample Exam', 'Quiz',
    }
    ws_exam = wb.create_sheet("Exams & Quizzes")
    style_header(ws_exam, [
        "#", "Course Code", "Document Name", "Document Type",
        "Semester / Year", "Has Key?", "Professor", "File Type", "Pages", "URL",
    ], use_blue=False)

    exam_docs = [d for d in docs_sorted if d['doc_type'] in EXAM_TYPES]
    for i, doc in enumerate(exam_docs, 1):
        ws_exam.append([
            i,
            doc['course_label'],
            doc['title'],
            doc['doc_type'],
            doc['semester_year'],
            'Yes' if doc['has_key'] else '',
            doc['professor'],
            doc['file_ext'].upper() if doc['file_ext'] else 'PDF',
            doc['pages'],
            doc['url'],
        ])
        style_data_row(ws_exam, i)

    ws_exam.freeze_panes = "A2"
    ws_exam.auto_filter.ref = ws_exam.dimensions
    for col, w in zip('ABCDEFGHIJ', [5, 14, 60, 28, 16, 10, 20, 10, 8, 72]):
        ws_exam.column_dimensions[col].width = w

    # ── Sheet: Instructional Material & Worksheets ────────────────────────
    MATERIAL_TYPES = {
        'Handout / Worksheet', 'Instructional Material',
        'Notes / Lecture Slides', 'Formula Sheet / Cheat Sheet',
        'Summary / Outline', 'Textbook Chapter / Reading',
    }
    ws_mat = wb.create_sheet("Instructional Material")
    style_header(ws_mat, [
        "#", "Course Code", "Document Name", "Document Type",
        "Semester / Year", "Professor", "File Type", "Pages", "URL",
    ], use_blue=False)

    mat_docs = [d for d in docs_sorted if d['doc_type'] in MATERIAL_TYPES]
    for i, doc in enumerate(mat_docs, 1):
        ws_mat.append([
            i,
            doc['course_label'],
            doc['title'],
            doc['doc_type'],
            doc['semester_year'],
            doc['professor'],
            doc['file_ext'].upper() if doc['file_ext'] else 'PDF',
            doc['pages'],
            doc['url'],
        ])
        style_data_row(ws_mat, i)

    ws_mat.freeze_panes = "A2"
    ws_mat.auto_filter.ref = ws_mat.dimensions
    for col, w in zip('ABCDEFGHI', [5, 14, 60, 28, 16, 20, 10, 8, 72]):
        ws_mat.column_dimensions[col].width = w

    # ── Sheet: Lab Reports ────────────────────────────────────────────────
    ws_lab = wb.create_sheet("Lab Reports")
    style_header(ws_lab, [
        "#", "Course Code", "Document Name", "Document Type",
        "Semester / Year", "Professor", "File Type", "Pages", "URL",
    ], use_blue=False)

    lab_docs = [d for d in docs_sorted if d['doc_type'] == 'Lab Report / Lab Manual']
    for i, doc in enumerate(lab_docs, 1):
        ws_lab.append([
            i,
            doc['course_label'],
            doc['title'],
            doc['doc_type'],
            doc['semester_year'],
            doc['professor'],
            doc['file_ext'].upper() if doc['file_ext'] else 'PDF',
            doc['pages'],
            doc['url'],
        ])
        style_data_row(ws_lab, i)

    ws_lab.freeze_panes = "A2"
    ws_lab.auto_filter.ref = ws_lab.dimensions
    for col, w in zip('ABCDEFGHI', [5, 14, 60, 28, 16, 20, 10, 8, 72]):
        ws_lab.column_dimensions[col].width = w

    # ── Sheet: Summary ────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.column_dimensions['A'].width = 42
    ws_sum.column_dimensions['B'].width = 80

    def sum_row(label, value='', bold=False, fill_color=None, font_color="000000"):
        ws_sum.append([str(label), str(value) if value != '' else ''])
        row = ws_sum.max_row
        ws_sum.row_dimensions[row].height = 18
        ws_sum[f'A{row}'].font = Font(bold=bold, size=10, color=font_color)
        ws_sum[f'B{row}'].font = Font(size=10, color=font_color)
        ws_sum[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws_sum[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        if fill_color:
            fl = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            ws_sum[f'A{row}'].fill = fl
            ws_sum[f'B{row}'].fill = fl

    # Title row
    sum_row("UCSB — Chemistry 1C CourseHero Document Inventory", "", bold=True,
            fill_color=UCSB_BLUE, font_color="FFFFFF")
    ws_sum.row_dimensions[1].height = 28

    ws_sum.append([])
    sum_row("School",     "University of California Santa Barbara", bold=True)
    sum_row("Courses",    "CHEM 1C / CHEM 1CL / CHEM-1C1C / CHEMISTRY 1C", bold=True)
    sum_row("Search Date", DATE_STAMP, bold=True)
    sum_row("Data File",  os.path.basename(json_file), bold=True)
    sum_row("Total Unique Documents", len(docs), bold=True)
    ws_sum.append([])

    # Source URLs with status
    sum_row("Course Sitemap URLs Searched (Status)", "", bold=True, fill_color=UCSB_CREAM, font_color=UCSB_BLUE)
    SOURCE_URLS = [
        ("CHEM 1C",       "https://www.coursehero.com/sitemap/schools/27-University-of-California-Santa-Barbara/courses/242694-CHEM1C/"),
        ("CHEM 1CL",      "https://www.coursehero.com/sitemap/schools/27-University-of-California-Santa-Barbara/courses/242693-CHEM1CL/"),
        ("CHEM-1C1C",     "https://www.coursehero.com/sitemap/schools/27-University-of-California-Santa-Barbara/courses/906488-CHEM-1C1C/"),
        ("CHEMISTRY 1C",  "https://www.coursehero.com/sitemap/schools/27-University-of-California-Santa-Barbara/courses/4576471-CHEMISTRY1C/"),
    ]
    for label, url in SOURCE_URLS:
        status_info = url_status.get(url, {"status": "unknown", "error": ""})
        status = status_info.get('status', 'unknown')
        error = status_info.get('error', '')
        
        if status == 'success':
            status_label = f"✓ SUCCESS"
        elif status == 'cloudflare':
            status_label = f"✗ CLOUDFLARE BLOCK"
        elif status == 'timeout':
            status_label = f"✗ TIMEOUT"
        elif status == 'wifi':
            status_label = f"✗ WIFI/CONNECTION LOSS"
        else:
            status_label = f"? {status.upper()}"
        
        display_val = f"{label}: {status_label}"
        if error:
            display_val += f" ({error[:40]}...)" if len(error) > 40 else f" ({error})"
        
        sum_row(display_val, url)
    ws_sum.append([])

    # Per-course breakdown
    sum_row("Documents per Course", "", bold=True, fill_color=UCSB_CREAM, font_color=UCSB_BLUE)
    course_counts = Counter(d['course_label'] for d in docs)
    ws_sum.append(["Course", "Count"])
    ws_sum[f'A{ws_sum.max_row}'].font = Font(bold=True, size=10)
    ws_sum[f'B{ws_sum.max_row}'].font = Font(bold=True, size=10)
    for label, _ in SOURCE_URLS:
        for course, count in sorted(course_counts.items()):
            if label in course or course in label:
                sum_row(course, count)
                break
    ws_sum.append([])

    # Document type breakdown
    type_counts = Counter(d['doc_type'] for d in docs)
    sum_row("Documents by Type", "", bold=True, fill_color=UCSB_CREAM, font_color=UCSB_BLUE)
    ws_sum.append(["Document Type", "Count"])
    ws_sum[f'A{ws_sum.max_row}'].font = Font(bold=True, size=10)
    ws_sum[f'B{ws_sum.max_row}'].font = Font(bold=True, size=10)
    for t in TYPE_ORDER:
        if t in type_counts:
            sum_row(t, type_counts[t])
    for t, ct in type_counts.items():
        if t not in TYPE_ORDER:
            sum_row(t, ct)
    ws_sum.append([])

    # Exam-only counts
    exam_count = sum(1 for d in docs if d['doc_type'] in EXAM_TYPES)
    sum_row("Exams & Quizzes Only", "", bold=True, fill_color=UCSB_CREAM, font_color=UCSB_BLUE)
    sum_row("Total Exam/Quiz Documents", exam_count, bold=True)

    mat_count = sum(1 for d in docs if d['doc_type'] in MATERIAL_TYPES)
    sum_row("Total Instructional Material", mat_count, bold=True)
    lab_count = sum(1 for d in docs if d['doc_type'] == 'Lab Report / Lab Manual')
    sum_row("Total Lab Reports / Lab Manuals", lab_count, bold=True)
    ws_sum.append([])

    # Notes
    sum_row("Notes & Methodology", "", bold=True, fill_color=UCSB_BLUE, font_color="FFFFFF")
    notes = [
        "All document links sourced from live CourseHero course sitemap pages (paginated ?p=1, ?p=2, ...).",
        "Chrome CDP (DevTools Protocol) attach mode used to bypass Cloudflare/bot detection.",
        "Uploader profile scraping intentionally skipped to avoid CloudFlare rate-limit blocks.",
        "Document type classification is rule-based from the document file name/slug.",
        "Professor column populated only where a known UCSB Chemistry professor name appears in the file name.",
        "Semester/Year parsed from sitemap page metadata.",
        "All course code variants (CHEM, CHEMISTRY) scraped; documents de-duplicated by URL.",
        "File URLs follow the pattern: https://www.coursehero.com/file/<id>/<slug>/",
        "Documents requiring CourseHero login to view full content; URLs are direct-link verified.",
        "URL status tracked: SUCCESS (fully scraped), CLOUDFLARE BLOCK (429 error), TIMEOUT, or WIFI/CONNECTION LOSS.",
    ]
    for note in notes:
        sum_row("", note)

    wb.save(output_path)
    return docs, type_counts, exam_count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    json_file = find_json_file()
    if not json_file or not os.path.exists(json_file):
        print("ERROR: No JSON data file found.")
        print("Run scrape_ucsb_chem1c_with_retry.py first to generate the data file.")
        sys.exit(1)

    print(f"Loading data from: {json_file}")
    docs = build_docs(json_file)
    url_status = load_url_status(json_file)
    print(f"Loaded {len(docs)} unique documents.")

    out_path = os.path.join(
        SCRIPT_DIR,
        f"coursehero_UCSB_CHEM1C_CHEM1CL_CHEM-1C1C_CHEMISTRY1C_{DATE_STAMP}.xlsx"
    )
    docs_final, type_counts, exam_count = write_xlsx(docs, json_file, out_path, url_status)
    print(f"\nXLSX saved → {out_path}")
    print(f"  Total documents : {len(docs_final)}")
    print(f"  Exams & Quizzes : {exam_count}")
    print("\nDocument type breakdown:")
    for t in TYPE_ORDER:
        if t in type_counts:
            print(f"  {t}: {type_counts[t]}")


if __name__ == '__main__':
    main()
