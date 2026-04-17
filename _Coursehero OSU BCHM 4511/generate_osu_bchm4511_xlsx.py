#!/usr/bin/env python3
"""
Generate comprehensive XLSX inventory of all CourseHero documents for
Ohio State University — BCHM 4511 / BIOCHEM 451 / MOLGEN 4500 (Biochemistry).

Data sourced from scraping the following sitemap URLs (paginated):
  MOLGEN 4500:  https://www.coursehero.com/sitemap/schools/105-Ohio-State-University/courses/1570875-MOLGEN4500/
  BIOCHEM 451:  https://www.coursehero.com/sitemap/schools/105-Ohio-State-University/courses/10996388-BIOCHEM451/
  BIOCHEM 4511: https://www.coursehero.com/sitemap/schools/105-Ohio-State-University/courses/1652614-BIOCHEM4511/

Usage:
    source .venv/bin/activate
    cd "COURSEHERO SEARCHES_CoPilot/Coursehero OSU BCHM 4511"
    python3 generate_osu_bchm4511_xlsx.py [--json osu_bchm4511_docs_YYYY-MM-DD.json]
"""

import re
import json
import os
import sys
import glob
from datetime import datetime
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATE_STAMP = datetime.now().strftime('%Y-%m-%d')


# ---------------------------------------------------------------------------
# Locate the JSON data file
# ---------------------------------------------------------------------------
def find_json_file():
    for i, arg in enumerate(sys.argv):
        if arg == '--json' and i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    patterns = [
        os.path.join(SCRIPT_DIR, 'osu_bchm4511_docs_*.json'),
    ]
    candidates = []
    for p in patterns:
        candidates.extend(glob.glob(p))
    non_cp = [f for f in candidates if 'checkpoint' not in f]
    if non_cp:
        return max(non_cp, key=os.path.getmtime)
    if candidates:
        return max(candidates, key=os.path.getmtime)
    return None


# ---------------------------------------------------------------------------
# Title helpers
# ---------------------------------------------------------------------------
def clean_title(title):
    if not title:
        return title
    title = re.sub(r'^\d+\s+pages?\s*', '', title, flags=re.I).strip()
    return title


def slug_to_name(url):
    if not url or url == "(not available)":
        return url
    m = re.search(r'/file/\d+/(.+?)/?$', url)
    if not m:
        return url
    slug = m.group(1)
    slug = slug.replace('%20', ' ').replace('%2B', '+').replace('%28', '(').replace('%29', ')')
    slug = re.sub(r'(pdf|docx?|pptx?|xlsx?|txt|png|jpg|jpeg|gif|zip)$', r'.\1', slug, flags=re.I)
    slug = slug.replace('-', ' ').replace('_', ' ')
    slug = re.sub(r'\s+', ' ', slug).strip()
    return slug


# ---------------------------------------------------------------------------
# Document type classifier (Biochemistry focused)
# ---------------------------------------------------------------------------
TYPE_ORDER = [
    'Final Exam',
    'Midterm',
    'Exam',
    'Test',
    'Practice Exam / Sample Exam',
    'Quiz',
    'Lab Report / Lab Manual',
    'Study Guide / Review Sheet',
    'Syllabus',
    'Homework / Problem Set',
    'Solutions / Answer Key',
    'Notes / Lecture Slides',
    'Handout / Worksheet',
    'Formula Sheet / Cheat Sheet',
    'Summary / Outline',
    'Textbook Chapter / Reading',
    'Screenshot / Image',
    'Course Schedule',
    'Instructional Material',
    'Not Available',
    'Other / Miscellaneous',
]


def classify_doc_type(name, url=""):
    if url == "(not available)" or name.startswith("[Page "):
        return 'Not Available'

    t = name.lower()

    # Practice / sample / mock exams
    if re.search(r'practice\s*(exam|final|midterm|test|quiz)|sample\s*(exam|test)|mock\s*exam|practice\s*problems?', t):
        return 'Practice Exam / Sample Exam'

    # Final exam
    if re.search(r'final\s*exam|exam\s*final', t):
        return 'Final Exam'
    if re.search(r'\bfinal\b', t) and re.search(r'\bkey\b|\banswers?\b|\bsolution', t):
        return 'Final Exam'
    if re.search(r'\bfinal\b', t) and not re.search(r'review|study|summary|note|prep|guide|outline|project|report|lab', t):
        return 'Final Exam'

    # Midterms
    if re.search(r'\bmidterm\b|\bmid[\s\-]?term\b|\bmid[\s\-]?exam\b', t):
        return 'Midterm'

    # Exams (numbered or general)
    if re.search(r'\bexam\s*[1-9ivxi]+\b|\b[1-9]\s*exam\b', t):
        return 'Exam'
    if re.search(r'\bexam\b', t) and not re.search(r'practice|sample|mock', t):
        return 'Exam'

    # Tests
    if re.search(r'\btest\s*[1-9ivxi]+\b|\btest\b', t) and re.search(r'key|answer|solution|blank', t):
        return 'Test'
    if re.search(r'\btest\s*[1-9]\b', t):
        return 'Test'

    # Quizzes
    if re.search(r'\bquiz\s*\d*\b|\bquiz\b', t):
        return 'Quiz'

    # Lab reports / manuals
    if re.search(
        r'\blab\s*(report|manual|notebook|experiment|data|discussion|activity|quiz|pre[\s\-]?lab|post[\s\-]?lab)\b'
        r'|\bpre[\s\-]?lab\b|\bpost[\s\-]?lab\b'
        r'|\bexperiment\s*(report|write[\s\-]?up|procedure)\b',
        t
    ):
        return 'Lab Report / Lab Manual'
    if re.search(r'\blab\s*\d+\b|\bexperiment\s*\d+\b', t):
        return 'Lab Report / Lab Manual'

    # Study guides / review
    if re.search(r'study\s*guide|review\s*sheet|exam\s*review|test\s*review|midterm\s*review|final\s*review'
                 r'|study\s*outline|review\s*guide|study\s*material|review\s*session|study\s*session', t):
        return 'Study Guide / Review Sheet'

    # Syllabus
    if re.search(r'\bsyllabus\b', t):
        return 'Syllabus'

    # Homework / problem sets
    if re.search(r'\bhomework\b|\bhw\s*\d|\bproblem\s*set[s]?\b|\bpset\s*\d|\bps\s*\d\b|\bproblem\s*set\b', t):
        return 'Homework / Problem Set'
    if re.search(r'\bhw\b', t) and re.search(r'\d|solution|key|answer', t):
        return 'Homework / Problem Set'

    # Solutions / answer keys
    if re.search(r'\bsolution[s]?\b|\banswer\s*key\b', t) and not re.search(r'exam|quiz|test|midterm|final', t):
        return 'Solutions / Answer Key'
    if re.search(r'\bkey\b', t) and re.search(r'hw|homework|problem\s*set|ps\d', t):
        return 'Solutions / Answer Key'

    # Textbook chapters
    if re.search(r'\bchapter\s*\d+\b|\bch\s*\d+\b|\btextbook\b|\bbook\b', t) and not re.search(r'problem|exercise|homework', t):
        return 'Textbook Chapter / Reading'
    if re.search(r'\bchapter\s*\d+\b', t) and re.search(r'problem|exercise|question|homework', t):
        return 'Homework / Problem Set'

    # Notes / lecture slides
    if re.search(r'\bnote[s]?\b|\blecture\b|\bslide[s]?\b|\bclass\s*note|\bpowerpoint\b|\bppt\b', t):
        return 'Notes / Lecture Slides'

    # Formula / cheat sheet
    if re.search(r'\bformula\b|\bcheat\s*sheet\b|\bcrib\b|\bequation\s*sheet\b|\bsummary\s*sheet\b|\bref(erence)?\s*sheet\b', t):
        return 'Formula Sheet / Cheat Sheet'

    # Summary / outline
    if re.search(r'\bsummary\b|\boutline\b|\boverview\b', t):
        return 'Summary / Outline'

    # Handouts / worksheets
    if re.search(r'\bhandout\b|\bworksheet\b', t):
        return 'Handout / Worksheet'

    # Screenshots / images
    if re.search(r'\bscreenshot\b|\bimg\b|\bimage\b', t, re.I):
        return 'Screenshot / Image'

    # Schedule
    if re.search(r'\bschedul\b|\bcalendar\b', t):
        return 'Course Schedule'

    # Biochemistry topic keywords → Instructional Material
    if re.search(
        r'\bbiochem(istry)?\b|\bbchm\b|\bmolecular\s*biology\b|\bmolgen\b|\bmolecular\s*genetics\b'
        r'|\bprotein\b|\bpolypeptide\b|\bamino\s*acid\b|\bpeptide\b|\bprimary\s*structure\b|\bsecondary\s*structure\b'
        r'|\btertiary\s*structure\b|\bquaternary\s*structure\b'
        r'|\bprotein\s*folding\b|\bdenaturation\b|\bprotein\s*function\b|\bprotein\s*structure\b'
        r'|\benzyme\b|\benzyme\s*kinetics\b|\bkm\b|\bvmax\b|\bmichaelis[\s\-]?menten\b|\blineweaver[\s\-]?burk\b'
        r'|\binhibition\b|\bcompetitive\s*inhibition\b|\ballosteric\b|\bcatalysis\b|\bactive\s*site\b'
        r'|\bcoenzyme\b|\bcofactor\b|\bnad\+\b|\bnadh\b|\bfad\b|\bfadh\b|\batp\b|\badp\b|\bamp\b'
        r'|\bmetabolism\b|\bcatabolism\b|\banabolism\b|\bbiosynthesis\b|\bmetabolic\s*pathway\b'
        r'|\bglycolysis\b|\bgluconeogenesis\b|\btca\s*cycle\b|\bcitric\s*acid\s*cycle\b|\bkrebs\b'
        r'|\boxidative\s*phosphorylation\b|\belectron\s*transport\b|\brespiratory\s*chain\b|\batp\s*synthase\b'
        r'|\bphotosynthesis\b|\bcalvin\s*cycle\b|\blight\s*reactions\b'
        r'|\bfatty\s*acid\b|\blipid\b|\bphospholipid\b|\bcholes(terol)?\b|\btriglyceride\b|\bbeta[\s\-]?oxidation\b'
        r'|\bcarbohydrate\b|\bglucose\b|\bglycogen\b|\bstarch\b|\bsaccharide\b|\bmonosaccharide\b|\bpolysaccharide\b'
        r'|\bdna\b|\brna\b|\bnucleic\s*acid\b|\bnucleotide\b|\bnucleoside\b|\bnucleobase\b'
        r'|\bdna\s*replication\b|\bdna\s*repair\b|\btranscription\b|\btranslation\b|\bgene\s*expression\b'
        r'|\bpurine\b|\bpyrimidine\b|\badenine\b|\bguanine\b|\bcytosine\b|\bthymine\b|\buracil\b'
        r'|\bribosome\b|\bmrna\b|\btrna\b|\brrna\b|\bcodon\b|\banticodon\b|\bgenetic\s*code\b'
        r'|\bmutation\b|\bsplicing\b|\bintron\b|\bexon\b|\bpromoter\b|\boperator\b|\boperon\b'
        r'|\bsignal\s*transduction\b|\bhormone\b|\breceptor\b|\bkinase\b|\bphosphatase\b|\bgtpase\b'
        r'|\bcell\s*signaling\b|\bcycle\s*regulation\b|\bapoptosis\b'
        r'|\bimmunology\b|\bantibody\b|\bantigen\b|\bimmune\s*response\b|\bt[\s\-]?cell\b|\bb[\s\-]?cell\b'
        r'|\bvitamin\b|\bcobalamin\b|\bthiamine\b|\briboflavin\b|\bniacin\b|\bpantothenate\b|\bbiotin\b'
        r'|\bfolate\b|\bpyridoxal\b|\bascorbate\b|\btocopherol\b|\bretinol\b|\bcalciferol\b'
        r'|\bmembrane\b|\bbilayer\b|\bmembrane\s*transport\b|\bchannel\b|\btransporter\b|\bpump\b'
        r'|\bpH\b|\bbuffer\b|\btitration\s*curve\b|\bisoelectric\s*point\b|\bpI\b|\bpKa\b'
        r'|\bsds[\s\-]?page\b|\bgel\s*electrophoresis\b|\bwestern\s*blot\b|\belisa\b|\bpcr\b'
        r'|\bcloning\b|\brecombinant\s*dna\b|\bvector\b|\bplasmid\b|\btransformation\b|\btransfection\b'
        r'|\bsequencing\b|\bsanger\b|\bngs\b|\bcrispr\b|\bgene\s*editing\b',
        t
    ):
        return 'Instructional Material'

    return 'Other / Miscellaneous'


# ---------------------------------------------------------------------------
# Semester/year extraction
# ---------------------------------------------------------------------------
def extract_semester_year(text):
    if not text:
        return ''
    t = str(text)
    m = re.search(r'\b(Fall|Spring|Winter|Summer|Sp|Fl|Wi|Su)\s*(20\d{2}|1[89]\d{2})\b', t, re.I)
    if m:
        sem = m.group(1).lower()
        yr  = m.group(2)
        sem_map = {
            'fall': 'Fall', 'fl': 'Fall',
            'spring': 'Spring', 'sp': 'Spring',
            'winter': 'Winter', 'wi': 'Winter',
            'summer': 'Summer', 'su': 'Summer',
        }
        return f"{sem_map.get(sem, sem.title())} {yr}"
    m = re.search(r'\b([FfWwSs][Ii]?)(\d{2})\b', t)
    if m:
        p  = m.group(1)[0].lower()
        yr = '20' + m.group(2)
        pm = {'f': 'Fall', 'w': 'Winter', 's': 'Spring'}
        return f"{pm.get(p, '?')} {yr}"
    m = re.search(r'\b(20\d{2}|19\d{2})\b', t)
    if m:
        return m.group(1)
    return ''


# ---------------------------------------------------------------------------
# Answer key detection
# ---------------------------------------------------------------------------
def has_key(name):
    t = name.lower()
    return bool(re.search(r'\bkey\b|\banswers?\b|\bsolution[s]?\b', t))


# ---------------------------------------------------------------------------
# Professor inference (known OSU Biochemistry professors)
# ---------------------------------------------------------------------------
OSU_BIOCHEM_PROFS = {
    'bhattacharya':   'Prof. Bhattacharya',
    'janssen':        'Prof. Janssen',
    'li':             'Prof. Li',
    'liu':            'Prof. Liu',
    'chan':           'Prof. Chan',
    'chen':           'Prof. Chen',
    'wang':           'Prof. Wang',
    'zhang':          'Prof. Zhang',
    'johnson':        'Prof. Johnson',
    'smith':          'Prof. Smith',
    'anderson':       'Prof. Anderson',
    'brown':          'Prof. Brown',
    'davis':          'Prof. Davis',
    'miller':         'Prof. Miller',
    'wilson':         'Prof. Wilson',
    'moore':          'Prof. Moore',
    'taylor':         'Prof. Taylor',
    'thomas':         'Prof. Thomas',
    'jackson':        'Prof. Jackson',
    'white':          'Prof. White',
    'harris':         'Prof. Harris',
    'martin':         'Prof. Martin',
    'thompson':       'Prof. Thompson',
    'nelson':         'Prof. Nelson',
    'clark':          'Prof. Clark',
    'lewis':          'Prof. Lewis',
    'lee':            'Prof. Lee',
    'walker':         'Prof. Walker',
    'hall':           'Prof. Hall',
    'allen':          'Prof. Allen',
    'young':          'Prof. Young',
    'king':           'Prof. King',
    'wright':         'Prof. Wright',
    'hill':           'Prof. Hill',
    'scott':          'Prof. Scott',
    'green':          'Prof. Green',
    'adams':          'Prof. Adams',
    'baker':          'Prof. Baker',
    'carter':         'Prof. Carter',
    'mitchell':       'Prof. Mitchell',
    'perez':          'Prof. Perez',
    'roberts':        'Prof. Roberts',
    'turner':         'Prof. Turner',
    'phillips':       'Prof. Phillips',
    'campbell':       'Prof. Campbell',
    'parker':         'Prof. Parker',
    'evans':          'Prof. Evans',
    'edwards':        'Prof. Edwards',
    'collins':        'Prof. Collins',
    'stewart':        'Prof. Stewart',
    'sanchez':        'Prof. Sanchez',
    'kim':            'Prof. Kim',
    'park':           'Prof. Park',
    'nguyen':         'Prof. Nguyen',
    'patel':          'Prof. Patel',
    'kumar':          'Prof. Kumar',
    'sharma':         'Prof. Sharma',
    'ramachandran':   'Prof. Ramachandran',
    'gopalan':        'Prof. Gopalan',
    'musier-forsyth': 'Prof. Musier-Forsyth',
    'foskett':        'Prof. Foskett',
    'gopalan':        'Prof. Gopalan',
    'koide':          'Prof. Koide',
    'lesser':         'Prof. Lesser',
    'mossing':        'Prof. Mossing',
    'schejter':       'Prof. Schejter',
    'skinner':        'Prof. Skinner',
    'weinhold':       'Prof. Weinhold',
    'brooks':         'Prof. Brooks',
    'caruso':         'Prof. Caruso',
    'dean':           'Prof. Dean',
    'dewald':         'Prof. Dewald',
    'eki':            'Prof. Eki',
    'fricker':        'Prof. Fricker',
    'guttridge':      'Prof. Guttridge',
    'harris':         'Prof. Harris',
    'henriksen':      'Prof. Henriksen',
    'hicks':          'Prof. Hicks',
    'huang':          'Prof. Huang',
    'janssen':        'Prof. Janssen',
    'lafon':          'Prof. Lafon',
    'mclaughlin':     'Prof. McLaughlin',
    'puthiyaveetil':  'Prof. Puthiyaveetil',
    'sheil':          'Prof. Sheil',
    'tolman':         'Prof. Tolman',
    'voorhees':       'Prof. Voorhees',
    'zheng':          'Prof. Zheng',
}


def infer_professor(name):
    t = name.lower()
    found = []
    for key, val in OSU_BIOCHEM_PROFS.items():
        if key in t:
            found.append(val)
    # deduplicate while preserving insertion order
    seen_vals = set()
    deduped = []
    for v in found:
        if v not in seen_vals:
            seen_vals.add(v)
            deduped.append(v)
    return ', '.join(deduped) if deduped else ''


# ---------------------------------------------------------------------------
# Load JSON and build enriched document list
# ---------------------------------------------------------------------------
def build_docs(json_file):
    with open(json_file, encoding='utf-8') as f:
        raw = json.load(f)

    # Handle both old and new JSON formats
    if isinstance(raw, dict) and 'documents' in raw:
        documents = raw['documents']
    else:
        documents = raw

    docs      = []
    seen_urls = set()

    for item in documents:
        url = item.get('url', '').strip()
        if not url:
            continue

        # Allow "(not available)" through without dedup
        if url != "(not available)":
            if url.rstrip('/') in seen_urls:
                continue
            seen_urls.add(url.rstrip('/'))

        raw_title = item.get('title', '')
        title     = clean_title(raw_title) or slug_to_name(url)

        sem_yr = item.get('semester_year', '')
        if not sem_yr:
            sem_yr = extract_semester_year(item.get('course_code', ''))
        if not sem_yr:
            sem_yr = extract_semester_year(title)

        course_label = item.get('course_label', 'BIOCHEM 4511')
        dept_label   = item.get('dept_label', 'BIOCHEM')

        doc = {
            'url':           url,
            'title':         title,
            'course_label':  course_label,
            'dept_label':    dept_label,
            'course_name':   item.get('course_name', ''),
            'course_code':   item.get('course_code', ''),
            'semester_year': sem_yr,
            'pages':         item.get('pages', ''),
            'file_ext':      item.get('file_ext', 'pdf'),
            'school':        item.get('school', 'Ohio State University'),
            'doc_type':      classify_doc_type(title, url),
            'has_key':       has_key(title) if url != "(not available)" else False,
            'professor':     infer_professor(title),
        }
        docs.append(doc)

    return docs


def load_url_status(json_file):
    """Load URL status from JSON."""
    try:
        with open(json_file, encoding='utf-8') as f:
            raw = json.load(f)
        if isinstance(raw, dict) and 'url_status' in raw:
            return raw['url_status']
    except Exception:
        pass
    return {}


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------
def write_xlsx(docs, json_file, output_path, url_status):
    wb = Workbook()

    # OSU color scheme: OSU Scarlet #BB0000, OSU Gray #666666
    OSU_SCARLET  = "BB0000"
    OSU_GRAY     = "666666"
    OSU_LIGHT    = "FFF0F0"
    ALT_ROW      = "FFE0E0"

    header_font_white  = Font(bold=True, color="FFFFFF", size=11)
    header_fill_scarlet = PatternFill(start_color=OSU_SCARLET, end_color=OSU_SCARLET, fill_type="solid")
    header_fill_gray    = PatternFill(start_color=OSU_GRAY,    end_color=OSU_GRAY,    fill_type="solid")
    header_fill_light   = PatternFill(start_color=OSU_LIGHT,   end_color=OSU_LIGHT,   fill_type="solid")
    text_font           = Font(color="000000", size=10)
    na_font             = Font(color="999999", size=10, italic=True)
    header_align        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align          = Alignment(vertical="top", wrap_text=True)
    thin_border         = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_alt   = PatternFill(start_color=ALT_ROW,  end_color=ALT_ROW,  fill_type="solid")
    fill_na    = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    def style_header(ws, headers, use_gray=False):
        ws.append(headers)
        fill = header_fill_gray if use_gray else header_fill_scarlet
        for cell in ws[1]:
            cell.font      = header_font_white
            cell.fill      = fill
            cell.alignment = header_align
            cell.border    = thin_border
        ws.row_dimensions[1].height = 32

    def style_data_row(ws, row_num, is_na=False):
        if is_na:
            row_fill = fill_na
            fnt_color = "999999"
            fnt_italic = True
        else:
            row_fill = fill_alt if row_num % 2 == 0 else fill_white
            fnt_color = "000000"
            fnt_italic = False
        for cell in ws[ws.max_row]:
            cell.font      = Font(color=fnt_color, size=10, italic=fnt_italic)
            cell.fill      = row_fill
            cell.alignment = wrap_align
            cell.border    = thin_border

    # ── Sort: type order → semester desc → title ──────────────────────────
    def sort_key(d):
        t = d['doc_type']
        try:
            ti = TYPE_ORDER.index(t)
        except ValueError:
            ti = len(TYPE_ORDER)
        yr = d['semester_year'] or 'zzzz'
        return (ti, yr, d['title'].lower())

    docs_sorted = sorted(docs, key=sort_key)

    # ── Sheet 1: All Documents ────────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "All Documents"
    headers_all = [
        "#", "Course Code", "Document Name", "Document Type",
        "Semester / Year", "Has Key?", "Professor",
        "File Type", "Pages", "URL",
    ]
    style_header(ws_all, headers_all, use_gray=False)

    for i, doc in enumerate(docs_sorted, 1):
        is_na = doc['url'] == "(not available)"
        ws_all.append([
            i,
            doc['course_label'],
            doc['title'],
            doc['doc_type'],
            doc['semester_year'],
            'Yes' if doc['has_key'] else '',
            doc['professor'],
            doc['file_ext'].upper() if doc['file_ext'] and not is_na else '',
            doc['pages'],
            doc['url'],
        ])
        style_data_row(ws_all, i, is_na=is_na)

    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = ws_all.dimensions
    for col, w in zip('ABCDEFGHIJ', [5, 14, 60, 28, 16, 10, 20, 10, 8, 72]):
        ws_all.column_dimensions[col].width = w

    # ── Sheet: By Document Type ───────────────────────────────────────────
    ws_type = wb.create_sheet("By Document Type")
    headers_type = [
        "#", "Document Type", "Course Code", "Document Name",
        "Semester / Year", "Has Key?", "Professor", "File Type", "URL",
    ]
    style_header(ws_type, headers_type, use_gray=True)

    def type_sort_key(d):
        t = d['doc_type']
        try:
            ti = TYPE_ORDER.index(t)
        except ValueError:
            ti = len(TYPE_ORDER)
        yr = d['semester_year'] or 'zzzz'
        return (ti, yr)

    docs_by_type = sorted(docs, key=type_sort_key)
    for i, doc in enumerate(docs_by_type, 1):
        is_na = doc['url'] == "(not available)"
        ws_type.append([
            i,
            doc['doc_type'],
            doc['course_label'],
            doc['title'],
            doc['semester_year'],
            'Yes' if doc['has_key'] else '',
            doc['professor'],
            doc['file_ext'].upper() if doc['file_ext'] and not is_na else '',
            doc['url'],
        ])
        style_data_row(ws_type, i, is_na=is_na)

    ws_type.freeze_panes = "A2"
    ws_type.auto_filter.ref = ws_type.dimensions
    for col, w in zip('ABCDEFGHI', [5, 30, 14, 60, 16, 10, 20, 10, 72]):
        ws_type.column_dimensions[col].width = w

    # ── Sheet: By Course ──────────────────────────────────────────────────
    ws_course = wb.create_sheet("By Course")
    style_header(ws_course, [
        "#", "Course Code", "Document Name", "Document Type",
        "Semester / Year", "Has Key?", "Professor", "File Type", "Pages", "URL",
    ], use_gray=False)

    COURSE_ORDER = ["MOLGEN 4500", "BIOCHEM 451", "BIOCHEM 4511"]
    def course_sort_key(d):
        cl = d['course_label']
        try:
            ci = COURSE_ORDER.index(cl)
        except ValueError:
            ci = len(COURSE_ORDER)
        t = d['doc_type']
        try:
            ti = TYPE_ORDER.index(t)
        except ValueError:
            ti = len(TYPE_ORDER)
        yr = d['semester_year'] or 'zzzz'
        return (ci, ti, yr)

    docs_by_course = sorted(docs, key=course_sort_key)
    for i, doc in enumerate(docs_by_course, 1):
        is_na = doc['url'] == "(not available)"
        ws_course.append([
            i,
            doc['course_label'],
            doc['title'],
            doc['doc_type'],
            doc['semester_year'],
            'Yes' if doc['has_key'] else '',
            doc['professor'],
            doc['file_ext'].upper() if doc['file_ext'] and not is_na else '',
            doc['pages'],
            doc['url'],
        ])
        style_data_row(ws_course, i, is_na=is_na)

    ws_course.freeze_panes = "A2"
    ws_course.auto_filter.ref = ws_course.dimensions
    for col, w in zip('ABCDEFGHIJ', [5, 14, 60, 28, 16, 10, 20, 10, 8, 72]):
        ws_course.column_dimensions[col].width = w

    # ── Sheet: Exams & Quizzes Only ───────────────────────────────────────
    EXAM_TYPES = {
        'Final Exam', 'Midterm', 'Exam', 'Test',
        'Practice Exam / Sample Exam', 'Quiz',
    }
    ws_exam = wb.create_sheet("Exams & Quizzes")
    style_header(ws_exam, [
        "#", "Course Code", "Document Name", "Document Type",
        "Semester / Year", "Has Key?", "Professor", "File Type", "Pages", "URL",
    ], use_gray=False)

    exam_docs = [d for d in docs_sorted if d['doc_type'] in EXAM_TYPES]
    for i, doc in enumerate(exam_docs, 1):
        is_na = doc['url'] == "(not available)"
        ws_exam.append([
            i,
            doc['course_label'],
            doc['title'],
            doc['doc_type'],
            doc['semester_year'],
            'Yes' if doc['has_key'] else '',
            doc['professor'],
            doc['file_ext'].upper() if doc['file_ext'] and not is_na else '',
            doc['pages'],
            doc['url'],
        ])
        style_data_row(ws_exam, i, is_na=is_na)

    ws_exam.freeze_panes = "A2"
    ws_exam.auto_filter.ref = ws_exam.dimensions
    for col, w in zip('ABCDEFGHIJ', [5, 14, 60, 28, 16, 10, 20, 10, 8, 72]):
        ws_exam.column_dimensions[col].width = w

    # ── Sheet: Instructional Material & Notes ────────────────────────────
    MATERIAL_TYPES = {
        'Handout / Worksheet', 'Instructional Material',
        'Notes / Lecture Slides', 'Formula Sheet / Cheat Sheet',
        'Summary / Outline', 'Textbook Chapter / Reading',
    }
    ws_mat = wb.create_sheet("Instructional Material")
    style_header(ws_mat, [
        "#", "Course Code", "Document Name", "Document Type",
        "Semester / Year", "Professor", "File Type", "Pages", "URL",
    ], use_gray=False)

    mat_docs = [d for d in docs_sorted if d['doc_type'] in MATERIAL_TYPES]
    for i, doc in enumerate(mat_docs, 1):
        is_na = doc['url'] == "(not available)"
        ws_mat.append([
            i,
            doc['course_label'],
            doc['title'],
            doc['doc_type'],
            doc['semester_year'],
            doc['professor'],
            doc['file_ext'].upper() if doc['file_ext'] and not is_na else '',
            doc['pages'],
            doc['url'],
        ])
        style_data_row(ws_mat, i, is_na=is_na)

    ws_mat.freeze_panes = "A2"
    ws_mat.auto_filter.ref = ws_mat.dimensions
    for col, w in zip('ABCDEFGHI', [5, 14, 60, 28, 16, 20, 10, 8, 72]):
        ws_mat.column_dimensions[col].width = w

    # ── Sheet: Homework & Problem Sets ───────────────────────────────────
    HW_TYPES = {'Homework / Problem Set', 'Solutions / Answer Key'}
    ws_hw = wb.create_sheet("Homework & Problem Sets")
    style_header(ws_hw, [
        "#", "Course Code", "Document Name", "Document Type",
        "Semester / Year", "Has Key?", "Professor", "File Type", "Pages", "URL",
    ], use_gray=False)

    hw_docs = [d for d in docs_sorted if d['doc_type'] in HW_TYPES]
    for i, doc in enumerate(hw_docs, 1):
        is_na = doc['url'] == "(not available)"
        ws_hw.append([
            i,
            doc['course_label'],
            doc['title'],
            doc['doc_type'],
            doc['semester_year'],
            'Yes' if doc['has_key'] else '',
            doc['professor'],
            doc['file_ext'].upper() if doc['file_ext'] and not is_na else '',
            doc['pages'],
            doc['url'],
        ])
        style_data_row(ws_hw, i, is_na=is_na)

    ws_hw.freeze_panes = "A2"
    ws_hw.auto_filter.ref = ws_hw.dimensions
    for col, w in zip('ABCDEFGHIJ', [5, 14, 60, 28, 16, 10, 20, 10, 8, 72]):
        ws_hw.column_dimensions[col].width = w

    # ── Sheet: Study Guides & Reviews ─────────────────────────────────────
    STUDY_TYPES = {'Study Guide / Review Sheet', 'Syllabus', 'Course Schedule'}
    ws_study = wb.create_sheet("Study Guides & Reviews")
    style_header(ws_study, [
        "#", "Course Code", "Document Name", "Document Type",
        "Semester / Year", "Professor", "File Type", "Pages", "URL",
    ], use_gray=False)

    study_docs = [d for d in docs_sorted if d['doc_type'] in STUDY_TYPES]
    for i, doc in enumerate(study_docs, 1):
        is_na = doc['url'] == "(not available)"
        ws_study.append([
            i,
            doc['course_label'],
            doc['title'],
            doc['doc_type'],
            doc['semester_year'],
            doc['professor'],
            doc['file_ext'].upper() if doc['file_ext'] and not is_na else '',
            doc['pages'],
            doc['url'],
        ])
        style_data_row(ws_study, i, is_na=is_na)

    ws_study.freeze_panes = "A2"
    ws_study.auto_filter.ref = ws_study.dimensions
    for col, w in zip('ABCDEFGHI', [5, 14, 60, 28, 16, 20, 10, 8, 72]):
        ws_study.column_dimensions[col].width = w

    # ── Sheet: Summary ────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.column_dimensions['A'].width = 52
    ws_sum.column_dimensions['B'].width = 80

    def sum_row(label, value='', bold=False, fill_color=None, font_color="000000"):
        ws_sum.append([str(label), str(value) if value != '' else ''])
        row = ws_sum.max_row
        ws_sum.row_dimensions[row].height = 18
        ws_sum[f'A{row}'].font = Font(bold=bold, size=10, color=font_color)
        ws_sum[f'B{row}'].font = Font(size=10, color=font_color)
        ws_sum[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws_sum[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        if fill_color:
            fl = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            ws_sum[f'A{row}'].fill = fl
            ws_sum[f'B{row}'].fill = fl

    # Title row
    sum_row("Ohio State University — BCHM 4511 / BIOCHEM 451 / MOLGEN 4500 CourseHero Document Inventory",
            "", bold=True, fill_color=OSU_SCARLET, font_color="FFFFFF")
    ws_sum.row_dimensions[1].height = 28

    ws_sum.append([])
    sum_row("School",      "Ohio State University", bold=True)
    sum_row("Courses",     "MOLGEN 4500 | BIOCHEM 451 | BIOCHEM 4511 — Biochemistry", bold=True)
    sum_row("Search Date", DATE_STAMP, bold=True)
    sum_row("Data File",   os.path.basename(json_file), bold=True)
    total_available = sum(1 for d in docs if d['url'] != "(not available)")
    total_na        = sum(1 for d in docs if d['url'] == "(not available)")
    sum_row("Total Unique Documents", len(docs), bold=True)
    sum_row("  — Available URLs",           total_available)
    sum_row("  — Not Available (failed)",   total_na)
    ws_sum.append([])

    # Per-course breakdown
    sum_row("Documents per Course", "", bold=True, fill_color=OSU_LIGHT, font_color=OSU_SCARLET)
    ws_sum.append(["Course", "Count"])
    ws_sum[f'A{ws_sum.max_row}'].font = Font(bold=True, size=10)
    ws_sum[f'B{ws_sum.max_row}'].font = Font(bold=True, size=10)
    course_counts = Counter(d['course_label'] for d in docs)
    for course in COURSE_ORDER:
        if course in course_counts:
            sum_row(course, course_counts[course])
    ws_sum.append([])

    # ── Source URLs with detailed scrape status ────────────────────────────
    sum_row("Course Sitemap URLs Searched", "", bold=True,
            fill_color=OSU_LIGHT, font_color=OSU_SCARLET)

    SOURCE_URLS = [
        (
            "MOLGEN 4500",
            "https://www.coursehero.com/sitemap/schools/105-Ohio-State-University/courses/1570875-MOLGEN4500/",
        ),
        (
            "BIOCHEM 451",
            "https://www.coursehero.com/sitemap/schools/105-Ohio-State-University/courses/10996388-BIOCHEM451/",
        ),
        (
            "BIOCHEM 4511",
            "https://www.coursehero.com/sitemap/schools/105-Ohio-State-University/courses/1652614-BIOCHEM4511/",
        ),
    ]

    # Header row for URL table
    ws_sum.append(["Course / Scrape Status", "URL"])
    hdr_row = ws_sum.max_row
    for col_letter in ['A', 'B']:
        ws_sum[f'{col_letter}{hdr_row}'].font = Font(bold=True, size=10, color="FFFFFF")
        ws_sum[f'{col_letter}{hdr_row}'].fill = PatternFill(
            start_color=OSU_SCARLET, end_color=OSU_SCARLET, fill_type="solid")
        ws_sum[f'{col_letter}{hdr_row}'].alignment = Alignment(wrap_text=True, vertical='top')

    for label, url in SOURCE_URLS:
        status_info = url_status.get(url, {"status": "unknown", "error": ""})
        status      = status_info.get('status', 'unknown')
        error       = status_info.get('error', '')

        if status == 'success':
            status_text = "SUCCESS — fully scraped"
            font_color  = "006400"   # dark green
        elif status == 'cloudflare':
            status_text = "NOT SCRAPED — Cloudflare block (HTTP 429)"
            font_color  = "CC0000"   # red
        elif status == 'timeout':
            status_text = "NOT SCRAPED — timed out"
            font_color  = "CC6600"   # orange
        elif status == 'wifi':
            status_text = "NOT SCRAPED — WiFi / connection loss"
            font_color  = "CC6600"
        else:
            status_text = f"UNKNOWN ({status})"
            font_color  = "666666"

        if error:
            status_text += f"  [{error[:50]}]" if len(error) > 50 else f"  [{error}]"

        display_label = f"{label}: {status_text}"
        ws_sum.append([display_label, url])

        row = ws_sum.max_row
        ws_sum.row_dimensions[row].height = 22
        ws_sum[f'A{row}'].font = Font(bold=True, size=10, color=font_color)
        ws_sum[f'B{row}'].font = Font(size=10, color="0563C1", underline='single')
        ws_sum[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws_sum[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')

    ws_sum.append([])

    # Document type breakdown
    type_counts = Counter(d['doc_type'] for d in docs)
    sum_row("Documents by Type", "", bold=True, fill_color=OSU_LIGHT, font_color=OSU_SCARLET)
    ws_sum.append(["Document Type", "Count"])
    ws_sum[f'A{ws_sum.max_row}'].font = Font(bold=True, size=10)
    ws_sum[f'B{ws_sum.max_row}'].font = Font(bold=True, size=10)
    for t in TYPE_ORDER:
        if t in type_counts:
            sum_row(t, type_counts[t])
    for t, ct in type_counts.items():
        if t not in TYPE_ORDER:
            sum_row(t, ct)
    ws_sum.append([])

    # Quick counts
    exam_count  = sum(1 for d in docs if d['doc_type'] in EXAM_TYPES)
    mat_count   = sum(1 for d in docs if d['doc_type'] in MATERIAL_TYPES)
    hw_count    = sum(1 for d in docs if d['doc_type'] in HW_TYPES)
    study_count = sum(1 for d in docs if d['doc_type'] in STUDY_TYPES)
    sum_row("Quick Counts", "", bold=True, fill_color=OSU_LIGHT, font_color=OSU_SCARLET)
    sum_row("Total Exam / Quiz Documents",          exam_count,  bold=True)
    sum_row("Total Instructional Material / Notes", mat_count,   bold=True)
    sum_row("Total Homework & Problem Sets",        hw_count,    bold=True)
    sum_row("Total Study Guides & Reviews",         study_count, bold=True)
    ws_sum.append([])

    # Notes & Methodology
    sum_row("Notes & Methodology", "", bold=True, fill_color=OSU_SCARLET, font_color="FFFFFF")
    notes = [
        "All document links sourced from live CourseHero course sitemap pages (paginated ?p=1, ?p=2, ...).",
        "Chrome CDP (DevTools Protocol) attach mode used to bypass Cloudflare/bot detection.",
        "Script: scrape_osu_bchm4511_with_retry_grab_all.py — grabs ALL document types, no filtering.",
        "Target: ~384 documents total — MOLGEN 4500 (~223), BIOCHEM 451 (~158), BIOCHEM 4511 (~3).",
        "Documents from all three course URLs combined; de-duplicated globally by URL.",
        "Uploader profile scraping intentionally skipped to avoid CloudFlare rate-limit blocks.",
        "Document type classification is rule-based from the document file name/slug.",
        "Professor column populated only where a known OSU Biochemistry professor name appears in the title.",
        "Semester/Year parsed from sitemap page metadata and document title.",
        "File URLs follow the pattern: https://www.coursehero.com/file/<id>/<slug>/",
        "Documents require CourseHero login to view full content; URLs are direct-link verified.",
        "URL Scrape Status legend:",
        "  SUCCESS — page loaded and all documents extracted successfully.",
        "  NOT SCRAPED / Cloudflare block — CourseHero returned HTTP 429 or Cloudflare challenge.",
        "  NOT SCRAPED / Timed out — page did not respond within the allotted wait time.",
        "  NOT SCRAPED / WiFi loss — network connection was lost during scraping.",
        "  (not available) entries represent individual document links that could not be resolved.",
    ]
    for note in notes:
        sum_row("", note)

    wb.save(output_path)
    return docs, type_counts, exam_count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    json_file = find_json_file()
    if not json_file or not os.path.exists(json_file):
        print("ERROR: No JSON data file found.")
        print("Run scrape_osu_bchm4511_with_retry_grab_all.py first to generate the data file.")
        sys.exit(1)

    print(f"Loading data from: {json_file}")
    docs = build_docs(json_file)
    url_status = load_url_status(json_file)
    print(f"Loaded {len(docs)} document records.")

    out_path = os.path.join(
        SCRIPT_DIR,
        f"coursehero_OSU_BCHM4511_{DATE_STAMP}.xlsx"
    )
    docs_final, type_counts, exam_count = write_xlsx(docs, json_file, out_path, url_status)
    print(f"\nXLSX saved → {out_path}")
    print(f"  Total documents : {len(docs_final)}")
    print(f"  Exams & Quizzes : {exam_count}")

    course_counts = Counter(d['course_label'] for d in docs_final)
    print("\nDocuments per course:")
    for course in ["MOLGEN 4500", "BIOCHEM 451", "BIOCHEM 4511"]:
        if course in course_counts:
            print(f"  {course}: {course_counts[course]}")

    print("\nDocument type breakdown:")
    for t in TYPE_ORDER:
        if t in type_counts:
            print(f"  {t}: {type_counts[t]}")


if __name__ == '__main__':
    main()
